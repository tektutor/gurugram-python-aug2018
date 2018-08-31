#!/usr/bin/python

from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws['A1'] = 'Header 1' 
ws['B1'] = 'Header 2'
ws['C1'] = 'Header 3'

count = 0
index = 1

for row in ws.rows:
    if count == 0:
       count = 10
       break 
    for cell in row:
        cell.value = count * index
        index = index + 1
        ws.append( cell )

wb.save('test.xlsx')
