from openpyxl import load_workbook
wb = load_workbook(filename='test.xlsx', read_only=False)
ws = wb.active

ws['A1'] = ' '

wb.save('test.xlsx')
